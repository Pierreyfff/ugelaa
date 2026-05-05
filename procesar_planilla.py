import pandas as pd
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def extract_planilla_data(file_path):
    df = pd.read_excel(file_path, sheet_name=None)
    sheet = list(df.values())[0]
    
    rows = sheet.values.tolist()
    workers = []
    current_worker = None
    
    for i in range(len(rows)):
        row = rows[i]
        col0 = str(row[0]).strip() if pd.notna(row[0]) else ""
        col1 = row[1]
        col2 = str(row[2]).strip() if pd.notna(row[2]) else ""
        col3 = row[3]
        
        if col2 == "DETALLE" and col3 == "FEBRERO":
            if current_worker and "nombre" in current_worker:
                workers.append(current_worker)
            current_worker = {}
            
        elif col0 == "HABERES" and pd.notna(col1):
            current_worker["nombre"] = str(col1).strip()
            
        elif col0 == "HABERES" and col2 == "BASICA":
            current_worker["basica"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "PERSONAL":
            current_worker["personal"] = float(col3) if pd.notna(col3) else 0
            
        elif "DL" in col2 and len(col2) <= 7:
            current_worker["dl_asig"] = float(col3) if pd.notna(col3) else 60
            
        elif col2 == "TPH":
            current_worker["tph"] = float(col3) if pd.notna(col3) else 0
            
        elif pd.notna(col1) and col2 == "Diferencial":
            current_worker["cargo"] = str(col1).strip()
            current_worker["diferencial"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "Familia":
            current_worker["familia"] = float(col3) if pd.notna(col3) else 0
            
        elif pd.notna(col1) and col2 == "REF.MOV":
            current_worker["rd"] = str(col1).strip()
            current_worker["ref_mov"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "DS 021":
            current_worker["ds_021"] = float(col3) if pd.notna(col3) else 0
            
        elif pd.notna(col1) and col2 == "PREP. CLAS":
            current_worker["uu"] = str(col1).strip()
            current_worker["prep_clas"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "REUNIFICA":
            current_worker["reunifica"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "IGV":
            current_worker["igv"] = float(col3) if pd.notna(col3) else 0
            
        elif col0 == "DSCTOS" and col2 == "DL20530":
            current_worker["dl20530"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "SEG.SOC":
            current_worker["seg_soc"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "DRR. AD":
            current_worker["drr_ad"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "DRR. MAG":
            current_worker["drr_mag"] = float(col3) if pd.notna(col3) else 0
            
        elif "Desc. Judicial" in col2:
            current_worker["desc_judicial"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "COOP":
            current_worker["coop"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "FONAVI":
            current_worker["fonavi"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "Sindicato":
            current_worker["sindicato"] = float(col3) if pd.notna(col3) else 0
            
        elif col2 == "ptm":
            current_worker["ptm"] = float(col3) if pd.notna(col3) else 0
            
        elif col0 == "TOTAL HABERES":
            current_worker["total_haberes"] = float(col3) if pd.notna(col3) else 0
            
        elif col0 == "TOTAL DESCUENTOS":
            current_worker["total_descuentos"] = float(col3) if pd.notna(col3) else 0
            
        elif col0 == "TOTAL LIQUIDO":
            current_worker["total_liquido"] = float(col3) if pd.notna(col3) else 0
            
    if current_worker and "nombre" in current_worker:
        workers.append(current_worker)
    
    return workers

def generate_clean_excel(workers, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    headers = [
        "DNI", "Puesto", "RD", "UU",
        "BASICA", "PERSONAL", "DL_ASIG", "TPH", "DIFERENCIAL", "FAMILIA", 
        "REF_MOV", "DS_021", "PREP_CLAS", "REUNIFICA", "IGV",
        "DL20530", "SEG_SOC", "DRR_AD", "DRR_MAG", "COOP", "FONAVI", "SINDICATO", "PTM", "DESC_JUDICIAL",
        "TOTAL_HABERES", "TOTAL_DESCUENTOS", "TOTAL_LIQUIDO"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, w in enumerate(workers, 2):
        nombre = w.get("nombre", "")
        
        dni = ""
        if nombre:
            parts = nombre.split()
            for p in parts:
                if re.match(r'^\d{5,}$', p):
                    dni = p
                    break
        
        row_data = [
            dni,
            w.get("cargo", ""),
            w.get("rd", ""),
            w.get("uu", ""),
            w.get("basica", 0),
            w.get("personal", 0),
            w.get("dl_asig", 60),
            w.get("tph", 0),
            w.get("diferencial", 0),
            w.get("familia", 0),
            w.get("ref_mov", 0),
            w.get("ds_021", 0),
            w.get("prep_clas", 0),
            w.get("reunifica", 0),
            w.get("igv", 0),
            w.get("dl20530", 0),
            w.get("seg_soc", 0),
            w.get("drr_ad", 0),
            w.get("drr_mag", 0),
            w.get("coop", 0),
            w.get("fonavi", 0),
            w.get("sindicato", 0),
            w.get("ptm", 0),
            w.get("desc_judicial", 0),
            round(w.get("total_haberes", 0), 2),
            round(w.get("total_descuentos", 0), 2),
            round(w.get("total_liquido", 0), 2),
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14
    
    wb.save(output_path)
    print(f"Excel limpio guardado: {output_path}")

if __name__ == "__main__":
    input_file = "Constancia de haberes 1993.xls"
    output_file = "planilla_limpia.xlsx"
    
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    print(f"Procesando: {input_file}")
    workers = extract_planilla_data(input_file)
    
    print(f"\nEmpleados encontrados: {len(workers)}")
    for w in workers:
        print(f"  - {w.get('nombre', 'N/A')}: S/ {w.get('total_liquido', 0):.2f}")
    
    if workers:
        generate_clean_excel(workers, output_file)
        print("\nListo!")