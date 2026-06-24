import gc
import io
import json
import os
import re
import time
import requests
import xlrd
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app, origins=[
    "https://ugelaa.vercel.app",
    "http://localhost:5173",
    "http://localhost:3000",
])

BACKEND_URL = os.getenv("BACKEND_URL", "http://localhost:8080")
UPLOAD_FOLDER = "/app/uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

DNI_PATTERN = re.compile(r"DNI\s*(\d+)", re.IGNORECASE)
INST_KEYWORD = re.compile(r"INSTITUCION|INSTITUCIÓN", re.IGNORECASE)
DIST_KEYWORD = re.compile(r"DISTRITO", re.IGNORECASE)
IGNORAR_CONCEPTOS = {"REINTEGRO", "ESCOLARIDAD", "AGUINALDO", "DETALLE"}

# Patrones para identificar institución educativa
INST_PREFIX = re.compile(r'^(CE\b|CB\b|C\.E\.|E\.E\.|I\.E\.|IE\b|CEBA|CEBE|EE\s|EE\b|JARDIN|COLEGIO|ESCUELA|INSTITUCION|INSTITUTO)', re.IGNORECASE)
INST_NUMERO = re.compile(r'^\d{3,4}\s')

# Distritos conocidos de la provincia de Cañete
DISTRITOS_CONOCIDOS = [
    "ASIA", "CALANGO", "CERRO AZUL", "CHILCA", "COAYLLO", "IMPERIAL",
    "LUNAHUANÁ", "LUNAHUANA", "MALA", "NUEVO IMPERIAL", "PACARÁN",
    "QUILMANÁ", "QUILMANA", "SAN ANTONIO", "SAN LUIS",
    "SAN VICENTE DE CAÑETE", "SANTA CRUZ DE FLORES", "ZÚÑIGA", "ZUÑIGA"
]
DISTRITOS_SET = {d.upper() for d in DISTRITOS_CONOCIDOS}

# Palabras que NUNCA son distrito ni institución (son cargos u otros)
NO_INST_DIST = [
    "HABERES", "DSCTOS", "TOTAL", "DNI", "RD ", "R.D.", "RM ", "R.M.",
    "UU-", "PROF.", "DIRECTOR", "AUX.", "ASESOR", "SECRETARIA",
    "SECRETARIO", "OFICINISTA", "DOCENTE", "ESPECIALISTA",
    "COORDINADOR", "JEFE", "SUBDIRECTOR", "PROMOTOR", "TECNICO",
    "PSICOLOGO", "ENFERMERO", "MEDICO", "TRABAJADOR", "ADMINISTRATIVO",
    "BASICA", "DETALLE", "REINTEGRO", "ESCOLARIDAD", "AGUINALDO",
    "DL20", "DL25", "LEY ", "LEY-"
]

# Patrones para identificar cargos docentes
CARGO_PATRON = re.compile(
    r'^(PROF\.|PROF\s|DIRECTOR|AUX\.|ASESOR|SECRETARIA|SECRETARIO|'
    r'OFICINISTA|ESPECIALISTA|COORDINADOR|JEFE|SUBDIRECTOR|PROMOTOR|'
    r'TECNICO|PSICOLOGO|ENFERMERO|MEDICO|TRABAJADOR[A]?\s|'
    r'ADMINISTRATIVO|DOCENTE|INSPECTOR|BIBLIOTECARIO|LABORATORIO|'
    r'LABORAT\.|EDUCAC\.|EDUCACIÓN|EDUCACION)',
    re.IGNORECASE
)


def parsear_valor(v):
    if v is None:
        return None
    try:
        f = float(v)
        return round(f, 2) if f else None
    except (TypeError, ValueError):
        return None


def limpiar_concepto(c: str) -> str:
    return c.strip().lstrip("+").strip().upper()


def analizar_duplicados(empleados: list) -> dict:
    dni_count = {}
    nombre_count = {}
    dni_empleados = {}
    nombre_empleados = {}
    
    for idx, emp in enumerate(empleados):
        dni = emp.get("dni")
        nombre = emp.get("nombre")
        monto = emp.get("total_liquido")
        
        if dni:
            if dni not in dni_count:
                dni_count[dni] = 0
                dni_empleados[dni] = []
            dni_count[dni] += 1
            dni_empleados[dni].append({"idx": idx, "nombre": nombre, "monto": monto})
        
        if nombre:
            if nombre not in nombre_count:
                nombre_count[nombre] = 0
                nombre_empleados[nombre] = []
            nombre_count[nombre] += 1
            nombre_empleados[nombre].append({"idx": idx, "dni": dni, "monto": monto})
    
    dnis_duplicados = [(dni, count) for dni, count in dni_count.items() if count > 1]
    nombres_duplicados = [(nombre, count) for nombre, count in nombre_count.items() if count > 1]
    
    # Calculate exact duplicates: same DNI + same normalized name
    # These are the employees that will actually be skipped during import
    exactos_count = 0
    exactos_indices = []
    for dni, emp_list in dni_empleados.items():
        if len(emp_list) <= 1:
            continue
        name_groups = {}
        for emp in emp_list:
            nombre = emp.get("nombre") or ""
            norm_key = " ".join(nombre.strip().lower().split())
            if norm_key not in name_groups:
                name_groups[norm_key] = []
            name_groups[norm_key].append(emp)
        for name, group in name_groups.items():
            # Skip first occurrence, mark rest as exact duplicates
            for emp in group[1:]:
                exactos_count += 1
                exactos_indices.append(emp["idx"])
    
    return {
        "dnis_duplicados": [{"dni": dni, "count": count, "empleados": dni_empleados[dni]} for dni, count in dnis_duplicados],
        "nombres_duplicados": [{"nombre": nombre, "count": count, "empleados": nombre_empleados[nombre]} for nombre, count in nombres_duplicados],
        "exactos": exactos_count,
        "exactos_indices": exactos_indices,
    }


def calcular_monto_total(empleados: list) -> float:
    total = 0.0
    for emp in empleados:
        liquido = emp.get("total_liquido")
        if liquido:
            total += float(liquido)
    return round(total, 2)


def extraer_dni(texto: str):
    if not texto:
        return None
    m = DNI_PATTERN.search(str(texto))
    return m.group(1) if m else None


def _leer_filas(filepath: str) -> list:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xls":
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        filas = []
        for rx in range(ws.nrows):
            row = []
            for cx in range(ws.ncols):
                cell = ws.cell(rx, cx)
                if cell.ctype in (0, 5, 6):
                    row.append(None)
                elif cell.ctype == 2:
                    v = cell.value
                    row.append(int(v) if v == int(v) else v)
                else:
                    row.append(cell.value)
            filas.append(tuple(row))
        return filas
    else:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))
        wb.close()
        return filas


def _es_no_inst_dist(texto: str) -> bool:
    """Verifica si un texto definitivamente NO es institución ni distrito."""
    for palabra in NO_INST_DIST:
        if texto.upper().startswith(palabra):
            return True
    return False


def _extraer_inst(texto: str) -> str | None:
    """Extrae nombre de institución de un texto si coincide con patrones."""
    t = texto.strip().rstrip(":-–—")
    if _es_no_inst_dist(t):
        return None
    # Buscar keyword INSTITUCION
    m = INST_KEYWORD.search(t)
    if m:
        partes = INST_KEYWORD.split(t, maxsplit=1)
        if len(partes) > 1 and partes[1].strip():
            return partes[1].strip().rstrip(":-–—")
    # Buscar prefijos de institución educativa
    if INST_PREFIX.match(t):
        return t
    # Buscar números al inicio (ej: "2105 CENTRO DE VARONES")
    if INST_NUMERO.match(t):
        return t
    return None


def _extraer_distrito(texto: str) -> str | None:
    """Extrae nombre de distrito de un texto."""
    t = texto.strip().rstrip(":-–—").upper()
    if _es_no_inst_dist(t):
        return None
    # Buscar keyword DISTRITO
    m = DIST_KEYWORD.search(t)
    if m:
        partes = DIST_KEYWORD.split(t, maxsplit=1)
        if len(partes) > 1 and partes[1].strip():
            return partes[1].strip().rstrip(":-–—")
    # Buscar en lista de distritos conocidos (exacto)
    if t in DISTRITOS_SET:
        return t
    # Buscar coincidencia: texto contiene nombre de distrito
    for dist in DISTRITOS_CONOCIDOS:
        if dist.upper() in t:
            return dist
    # Buscar coincidencia parcial: texto es prefijo de distrito conocido
    for dist in DISTRITOS_CONOCIDOS:
        du = dist.upper()
        if du.startswith(t) and len(t) >= 3:
            return dist
    # Buscar coincidencia parcial invertida: texto contiene el inicio del distrito
    for dist in DISTRITOS_CONOCIDOS:
        du = dist.upper()
        primeros = du.split()[0] if ' ' in du else du
        if len(primeros) >= 4 and t.startswith(primeros):
            return dist
    return None


def escanear_encabezado(filas: list) -> dict:
    info = {"institucion": None, "distrito": None}
    primer_inst = None
    primer_dist = None

    for row in filas[:30]:
        # Solo revisamos columna B (index 1) y A (index 0) que es donde suelen ir
        col_a = str(row[0]).strip() if len(row) > 0 and row[0] else ""
        col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        col_c = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        for celda in (col_a, col_b, col_c):
            if not celda:
                continue
            texto = str(celda).strip()
            if not texto:
                continue

            # Detectar DISTRITO primero (por keyword o lista conocida)
            dist = _extraer_distrito(texto)
            if dist and info["distrito"] is None:
                info["distrito"] = dist
                primer_dist = texto

            # Detectar INSTITUCION
            inst = _extraer_inst(texto)
            if inst and info["institucion"] is None and texto != primer_dist:
                # Verificar que no sea un distrito que ya detectamos
                if _extraer_distrito(texto) is None:
                    info["institucion"] = inst
                    primer_inst = texto

        # Si ya tenemos ambos, podemos parar
        if info["institucion"] and info["distrito"]:
            break

    # Fallback: si solo tenemos institución, buscar distrito en filas
    # donde la columna B no coincide con institución
    if info["institucion"] and not info["distrito"]:
        for row in filas[:30]:
            col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if col_b and col_b != info["institucion"]:
                d = _extraer_distrito(col_b)
                if d:
                    info["distrito"] = d
                    break

    return info


def _clasificar_info_empleado(texto: str, emp: dict) -> None:
    """Clasifica un texto de columna B en DNI, RD, UU, cargo o institución/distrito."""
    if not texto:
        return
    t = texto.strip()

    # 1. Intentar extraer DNI
    dni = extraer_dni(texto)
    if dni and not emp["dni"]:
        emp["dni"] = dni
        return

    # 2. Detectar RD / Resolución
    if not emp["resolucion"] and re.match(
        r"^(RD|RM|DS|LEY|DL|R\.D\.|R\.M\.|R\.N\.|R\.G\.)\w*\s*[\d\-./A-Za-z\s]+",
        texto, re.IGNORECASE
    ):
        emp["resolucion"] = t
        return

    # 3. Detectar UU / Código
    if not emp["codigo"] and re.match(r"^uu-", texto, re.IGNORECASE):
        emp["codigo"] = t
        return

    # 4. Detectar CARGO docente por patrón
    if not emp["cargo"] and CARGO_PATRON.match(t):
        emp["cargo"] = t
        return

    # 5. Ignorar si es obviamente NO-institución/distrito
    texto_upper = t.upper()
    for no in NO_INST_DIST:
        if texto_upper.startswith(no):
            return

    # 6. Detectar DISTRITO (siempre intentar, solo si aún no tenemos)
    if not emp["distrito"]:
        dist = _extraer_distrito(texto)
        if dist:
            emp["distrito"] = dist
            return

    # 7. Detectar INSTITUCIÓN (siempre intentar, solo si aún no tenemos)
    if not emp["institucion"]:
        inst = _extraer_inst(texto)
        if inst and not _extraer_distrito(texto):
            emp["institucion"] = inst
            return

    # 8. CARGO fallback — solo si no es distrito ni institución
    if not emp["cargo"]:
        if not _extraer_distrito(texto) and not _extraer_inst(texto):
            emp["cargo"] = t


def extraer_empleados(filepath: str) -> list:
    filas = _leer_filas(filepath)
    encabezado = escanear_encabezado(filas)

    empleados = []
    i = 0
    n = len(filas)

    while i < n:
        fila = filas[i]
        col_a = str(fila[0]).strip() if fila[0] else ""

        if col_a == "HABERES" and len(fila) > 1 and fila[1]:
            nombre_completo = str(fila[1]).strip()

            # Buscar institución/distrito en las filas inmediatamente superiores
            # (el Excel tiene la institución en la columna B justo encima del nombre)
            inst_actual = encabezado["institucion"]
            dist_actual = encabezado["distrito"]
            j = i - 1
            while j >= 0 and j > i - 12:
                prev = filas[j]
                prev_a = str(prev[0]).strip() if len(prev) > 0 and prev[0] else ""
                prev_b = str(prev[1]).strip() if len(prev) > 1 and prev[1] else ""
                for celda in (prev_b, prev_a):
                    if not celda or celda == nombre_completo:
                        continue
                    inst = _extraer_inst(celda)
                    if inst and inst != nombre_completo:
                        inst_actual = inst
                    dist = _extraer_distrito(celda)
                    if dist and dist != nombre_completo:
                        dist_actual = dist
                if inst_actual and dist_actual:
                    break
                j -= 1

            emp = {
                "nombre": nombre_completo,
                "institucion": inst_actual,
                "distrito": dist_actual,
                "cargo": None,
                "resolucion": None,
                "codigo": None,
                "dni": None,
                "haberes": [],
                "descuentos": [],
                "total_haberes": None,
                "total_descuentos": None,
                "total_liquido": None,
            }

            # Columna C puede tener concepto inicial
            concepto_inicial = str(fila[2]).strip() if len(fila) > 2 and fila[2] else ""
            valor_inicial = parsear_valor(fila[3] if len(fila) > 3 else None)
            if concepto_inicial and limpiar_concepto(concepto_inicial) not in IGNORAR_CONCEPTOS:
                if valor_inicial is not None:
                    emp["haberes"].append({"concepto": concepto_inicial, "monto": valor_inicial})

            seccion = "HABERES"
            i += 1

            while i < n:
                f = filas[i]
                a = str(f[0]).strip() if len(f) > 0 and f[0] else ""
                b = str(f[1]).strip() if len(f) > 1 and f[1] else ""
                c_cell = str(f[2]).strip() if len(f) > 2 and f[2] else ""
                d = parsear_valor(f[3] if len(f) > 3 else None)

                # ── Totales ──
                if a == "TOTAL HABERES":
                    emp["total_haberes"] = d
                    i += 1
                    continue
                if a == "TOTAL DESCUENTOS":
                    emp["total_descuentos"] = d
                    i += 1
                    continue
                if a == "TOTAL LIQUIDO":
                    emp["total_liquido"] = d
                    i += 1
                    break

                # ── Sección DSCTOS ──
                if a == "DSCTOS":
                    seccion = "DSCTOS"
                    if c_cell and limpiar_concepto(c_cell) not in IGNORAR_CONCEPTOS and d is not None:
                        emp["descuentos"].append({"concepto": c_cell, "monto": d})
                    i += 1
                    continue

                # ── Nuevo bloque HABERES (siguiente empleado) ──
                if a == "HABERES" and len(f) > 1 and f[1]:
                    break

                # ── Procesar info del empleado en columna B ──
                if b:
                    _clasificar_info_empleado(b, emp)

                # ── Procesar conceptos (haberes/descuentos) en columna C ──
                if c_cell:
                    nombre_limpio = limpiar_concepto(c_cell)
                    if nombre_limpio not in IGNORAR_CONCEPTOS and not nombre_limpio.startswith("+"):
                        if d is not None:
                            item = {"concepto": c_cell.strip(), "monto": d}
                            if seccion == "HABERES":
                                emp["haberes"].append(item)
                            else:
                                emp["descuentos"].append(item)

                i += 1

            empleados.append(emp)
            continue

        i += 1

    return empleados


# ─────────────────────────────────────────────────────────────────────────────

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@app.route("/process-excel", methods=["POST"])
def process_excel():
    if "file" not in request.files:
        return jsonify({"error": "No se encontró archivo"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Archivo sin nombre"}), 400

    filename = secure_filename(file.filename)
    if not filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Solo se aceptan archivos Excel (.xlsx, .xls)"}), 400

    mes = request.form.get("mes", type=int)
    anio = request.form.get("anio", type=int)

    if not mes or not anio:
        return jsonify({"error": "Se requieren los campos mes y anio"}), 400
    if not (1 <= mes <= 12):
        return jsonify({"error": "Mes debe estar entre 1 y 12"}), 400

    edits_str = request.form.get("edits")
    edits = json.loads(edits_str) if edits_str else []

    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    file.save(filepath)

    try:
        empleados = extraer_empleados(filepath)

        # Apply user edits from pre-validation
        for edit in edits:
            idx = edit.get("idx")
            if idx is not None and 0 <= idx < len(empleados):
                if edit.get("dni"):
                    empleados[idx]["dni"] = edit["dni"]
                if edit.get("nombre"):
                    empleados[idx]["nombre"] = edit["nombre"]
        
        total_empleados = len(empleados)
        analisis = analizar_duplicados(empleados)
        monto_total = calcular_monto_total(empleados)

        # Pre-filter: remove exact duplicates (same DNI+name) before sending to Go
        exactos_indices_set = set(analisis["exactos_indices"])
        empleados_filtrados = [e for i, e in enumerate(empleados) if i not in exactos_indices_set]
        planillas_count = len(empleados_filtrados)
        dnis_dup = analisis["dnis_duplicados"]
        nombres_dup = analisis["nombres_duplicados"]
        exactos_count = analisis["exactos"]
        exactos_indices = analisis["exactos_indices"]

        del empleados, analisis
        gc.collect()

        payload = {
            "mes": mes,
            "anio": anio,
            "total_empleados": planillas_count,
            "empleados": empleados_filtrados,
        }

        response = requests.post(
            f"{BACKEND_URL}/api/importar/haberes",
            json=payload,
            timeout=300,
        )

        del empleados_filtrados, payload
        gc.collect()

        if response.status_code == 200:
            resp_data = response.json()
            return jsonify({
                "message": "Excel procesado correctamente",
                "personal_creados": resp_data.get("personal_creados", 0),
                "planillas_creadas": resp_data.get("planillas_creadas", 0),
                "personal": total_empleados,
                "planillas": planillas_count,
                "exactos": exactos_count,
                "errores": resp_data.get("errores", []),
                "personal_actualizados": resp_data.get("personal_actualizados", 0),
                "dnis_duplicados": dnis_dup,
                "nombres_duplicados": nombres_dup,
                "exactos_indices": exactos_indices,
                "monto_total": monto_total,
                "duplicados": resp_data.get("duplicados", []),
            })
        else:
            return jsonify({
                "error": "Error al enviar al backend",
                "details": response.text,
            }), response.status_code

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


@app.route("/validate-excel", methods=["POST"])
def validate_excel():
    if "file" not in request.files:
        return jsonify({"error": "No se encontró archivo"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Archivo sin nombre"}), 400

    filename = secure_filename(file.filename)
    if not filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Solo se aceptan archivos Excel (.xlsx, .xls)"}), 400

    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    file.save(filepath)

    try:
        empleados = extraer_empleados(filepath)
        total = len(empleados)
        preview = empleados[:5]
        analisis = analizar_duplicados(empleados)
        monto_total = calcular_monto_total(empleados)

        del empleados
        gc.collect()

        return jsonify({
            "valid": True,
            "total_empleados": total,
            "preview": preview,
            "dnis_duplicados": analisis["dnis_duplicados"],
            "nombres_duplicados": analisis["nombres_duplicados"],
            "exactos": analisis["exactos"],
            "exactos_indices": analisis["exactos_indices"],
            "monto_total": monto_total,
        })
    except Exception as e:
        return jsonify({"valid": False, "error": str(e)}), 400
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


MESES_ABR = ["", "ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SET", "OCT", "NOV", "DIC"]
MESES_NOMBRE = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

# Keywords that indicate a concept is a descuento, not a haber
_DESC_KEYWORDS = [
    "AFP", "ONP", "ESSALUD", "SNP", "COMISION", "SEGURO", "SALUD",
    "PRESTAMO", "PRÉSTAMO", "DESCUENTO", "DSCTO", "DESCTO",
    "RETENCION", "RETENCIÓN", "JUDICIAL", "SINDICATO",
    "MULTA", "CAF", "APORTE", "CUOTA", "FONAVI", "IMPUESTO",
    "RENTA", "QUINTA", "ADELANTO",
]


def _es_descuento(nombre: str) -> bool:
    """Return True if the concept name matches known descuento patterns."""
    upper = nombre.upper().replace(" ", "").replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    for kw in _DESC_KEYWORDS:
        if kw in upper:
            return True
    return False


from copy import copy

def _set_cell(ws, row, col, value):
    """Write safely: if target belongs to a merged range but is NOT the top‑left
    cell, unmerge the range first preserving formatting from the top‑left cell."""
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            if row != mr.min_row or col != mr.min_col:
                src = ws.cell(mr.min_row, mr.min_col)
                ws.unmerge_cells(
                    start_row=mr.min_row, start_column=mr.min_col,
                    end_row=mr.max_row, end_column=mr.max_col
                )
                fmt = dict(
                    border=copy(src.border),
                    font=copy(src.font),
                    alignment=copy(src.alignment),
                    number_format=copy(src.number_format),
                    fill=copy(src.fill),
                )
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        cell = ws.cell(r, c)
                        for attr, val in fmt.items():
                            setattr(cell, attr, val)
            break
    ws.cell(row=row, column=col, value=value)


def _unmerge_data_zone(ws):
    """Unmerge ALL merged cells in the data area (rows 9–40)."""
    for mr in list(ws.merged_cells.ranges):
        if 9 <= mr.min_row <= 40 or 9 <= mr.max_row <= 40:
            ws.unmerge_cells(
                start_row=mr.min_row, start_column=mr.min_col,
                end_row=mr.max_row, end_column=mr.max_col
            )


def _escribir_planilla_anual(ws, year, planillas_year, personal, logo_path):
    """Write one year's worth of data into a worksheet using the fixed template layout.

    Template layout (rows):
      9:  month abbreviations (D–O)
     10:  "HABERES" (col A) + month abbreviations cont.
     11–24: haberes concepts (max 14)
     25:  TOTAL HABERES  (per‑month in cols D–O)
     26:  empty
     27–34: descuentos concepts (max 8)
     35:  TOTAL DESCUENTOS (per‑month)
     36:  TOTAL LIQUIDO   (per‑month)
    Columns D=4=ENE … O=15=DIC
    """
    # ── Insert logo A1:C4 ────────────────────────────────────────────────
    ws["A1"].value = None
    if logo_path:
        try:
            from openpyxl.drawing.image import Image as XlImage
            img = XlImage(logo_path)
            img.width = 200
            img.height = 65
            ws.add_image(img, "A1")
        except Exception as ex:
            import sys
            print(f"[WARN] Logo load failed: {ex}", file=sys.stderr, flush=True)

    # ── Employee info (rows 5–8) ──────────────────────────────────────────
    _set_cell(ws, 5, 3,
              f"{personal.get('apellidos', '')}, {personal.get('nombres', '')}".strip(", "))
    _set_cell(ws, 7, 3, personal.get("dni", ""))
    _set_cell(ws, 5, 11, str(personal.get("institucion", "")))
    _set_cell(ws, 6, 11, str(personal.get("puesto", "")))
    _set_cell(ws, 7, 11, str(personal.get("rd", "")))
    _set_cell(ws, 8, 11, str(personal.get("uu", "")))

    # ── Period ────────────────────────────────────────────────────────────
    meses_con_datos = sorted(set(p["mes"] for p in planillas_year))
    if len(meses_con_datos) == 1:
        periodo_str = f"{MESES_NOMBRE[meses_con_datos[0]]} {year}"
    else:
        periodo_str = str(year)
    _set_cell(ws, 6, 3, periodo_str)

    # ── Month headers (row 9, col D–O) ────────────────────────────────────
    _set_cell(ws, 9, 1, "DESCRIPCION")
    for m in range(1, 13):
        _set_cell(ws, 9, 3 + m, MESES_ABR[m])

    # ── Section labels (row 10) ───────────────────────────────────────────
    _set_cell(ws, 10, 1, "HABERES")

    # ── Collect & classify concepts for this year ─────────────────────────
    haberes_data = {}   # concept_name -> {month: amount}
    descuentos_data = {}

    for p in planillas_year:
        try:
            mes = int(p["mes"])
        except (KeyError, TypeError, ValueError):
            continue
        # Ingress from DB → could be either haberes or descuentos depending
        # on how the original Excel was parsed. Re‑classify by concept name.
        for ing in p.get("ingresos", []):
            t = str(ing.get("tipo", "")).strip()
            if not t:
                continue
            try:
                monto = float(ing.get("monto", 0) or 0)
            except (TypeError, ValueError):
                monto = 0.0
            if _es_descuento(t):
                inner = descuentos_data.setdefault(t, {})
                inner[mes] = inner.get(mes, 0) + monto
            else:
                inner = haberes_data.setdefault(t, {})
                inner[mes] = inner.get(mes, 0) + monto

        # Descuentos from DB → always descuentos
        for desc in p.get("descuentos", []):
            t = str(desc.get("tipo", "")).strip()
            if not t:
                continue
            try:
                monto = float(desc.get("monto", 0) or 0)
            except (TypeError, ValueError):
                monto = 0.0
            inner = descuentos_data.setdefault(t, {})
            inner[mes] = inner.get(mes, 0) + monto

    # ── Write HABERES (rows 11–24, max 14) ────────────────────────────────
    haberes_sorted = sorted(haberes_data.keys())[:14]
    rh = 11
    for concepto in haberes_sorted:
        _set_cell(ws, rh, 1, str(concepto))
        for m in range(1, 13):
            _set_cell(ws, rh, 3 + m, haberes_data.get(concepto, {}).get(m, 0))
        rh += 1

    # ── TOTAL HABERES (row 25) ───────────────────────────────────────────
    _set_cell(ws, 25, 1, "TOTAL HABERES")
    for m in range(1, 13):
        total = sum(haberes_data.get(c, {}).get(m, 0) for c in haberes_sorted)
        _set_cell(ws, 25, 3 + m, round(total, 2))

    # row 26 intentionally left empty

    # ── Write DESCUENTOS (rows 26–34) ────────────────────────────────────
    # Row 26 = section header (template already has it)
    _set_cell(ws, 26, 1, "DESCUENTOS")
    # Rows 27–34 = concept rows (max 8)
    desc_sorted = sorted(descuentos_data.keys())[:8]
    for i, concepto in enumerate(desc_sorted):
        row = 27 + i
        _set_cell(ws, row, 1, str(concepto))
        for m in range(1, 13):
            _set_cell(ws, row, 3 + m, descuentos_data.get(concepto, {}).get(m, 0))

    # ── TOTAL DESCUENTOS (row 35) ────────────────────────────────────────
    _set_cell(ws, 35, 1, "TOTAL DESCUENTOS")
    for m in range(1, 13):
        total = sum(descuentos_data.get(c, {}).get(m, 0) for c in desc_sorted)
        _set_cell(ws, 35, 3 + m, round(total, 2))

    # ── TOTAL LIQUIDO (row 36) ───────────────────────────────────────────
    _set_cell(ws, 36, 1, "TOTAL LIQUIDO")
    for m in range(1, 13):
        th = sum(haberes_data.get(c, {}).get(m, 0) for c in haberes_sorted)
        td = sum(descuentos_data.get(c, {}).get(m, 0) for c in desc_sorted)
        _set_cell(ws, 36, 3 + m, round(th - td, 2))


@app.route("/export-excel", methods=["POST"])
def export_excel():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Se requiere JSON body"}), 400

    personal_id = data.get("personal_id")
    if not personal_id:
        return jsonify({"error": "Se requiere personal_id"}), 400

    params = {}
    if data.get("mes"):
        params["mes"] = str(data["mes"])
    if data.get("anio"):
        params["anio"] = str(data["anio"])

    try:
        resp = requests.get(
            f"{BACKEND_URL}/api/personal/{personal_id}/exportar",
            params=params,
            timeout=30,
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": f"Error de conexión con el backend: {str(e)}"}), 502

    if resp.status_code != 200:
        return jsonify({"error": "Error al obtener datos del empleado", "details": resp.text}), resp.status_code

    try:
        d = resp.json()
    except Exception as e:
        return jsonify({"error": f"Error al leer respuesta del backend: {str(e)}"}), 502

    personal = d.get("personal", {})
    planillas = d.get("planillas", [])

    # ── Locate template and logo ──────────────────────────────────────────
    base_dir = os.path.dirname(__file__)
    template_path = os.path.join(app.config["UPLOAD_FOLDER"], "plantilla_nueva.xlsx")
    if not os.path.exists(template_path):
        template_path = os.path.join(base_dir, "plantilla_nueva.xlsx")
    if not os.path.exists(template_path):
        return jsonify({"error": "Plantilla no encontrada en el servidor"}), 500

    logo_path = os.path.join(base_dir, "logo_minedu-.png")
    if not os.path.exists(logo_path):
        logo_path = None

    if not planillas:
        return jsonify({"error": "El empleado no tiene planillas registradas"}), 404

    # ── Group by year & sort ──────────────────────────────────────────────
    years = sorted(set(p["anio"] for p in planillas))

    try:
        from openpyxl import load_workbook
        wb = load_workbook(template_path)

        # Create one sheet per year
        template_ws = wb.active
        template_ws.title = str(years[0])
        year_sheets = {years[0]: template_ws}

        for yr in years[1:]:
            ws_copy = wb.copy_worksheet(template_ws)
            ws_copy.title = str(yr)
            year_sheets[yr] = ws_copy

        # Write each year's data
        for yr, ws in year_sheets.items():
            year_planillas = [p for p in planillas if p["anio"] == yr]
            _escribir_planilla_anual(ws, yr, year_planillas, personal, logo_path)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        ape = personal.get("apellidos", "").replace(" ", "_")
        nom = personal.get("nombres", "").replace(" ", "_")
        filename = f"Planilla_{ape}_{nom}.xlsx"

        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[EXPORT ERROR] {tb}", flush=True)
        return jsonify({"error": f"Error al generar el Excel: {repr(e)}\n{tb}"}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8081)
