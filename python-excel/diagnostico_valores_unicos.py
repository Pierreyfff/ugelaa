"""
Script de diagnóstico: extrae valores únicos por campo del Excel seleccionado.
Ayuda a identificar inconsistencias en cargos, distritos, instituciones, etc.
Uso:
    python diagnostico_valores_unicos.py
Seleccionar archivo: Constancia de haberes 1993_marzo.xlsx
"""
import os
import sys
import tkinter as tk
from tkinter import filedialog
from collections import Counter
from openpyxl import load_workbook


def seleccionar_archivo():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo Excel",
        filetypes=[("Archivos Excel", "*.xlsx *.xls"), ("Todos", "*.*")],
        initialdir=os.path.join(os.path.dirname(__file__), "uploads"),
    )
    root.destroy()
    return archivo if archivo else None


def leer_filas(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        filas = []
        for rx in range(ws.nrows):
            row = []
            for cx in range(ws.ncols):
                cell = ws.cell(rx, cx)
                if cell.ctype in (0, 5, 6):
                    row.append(None)
                elif cell.ctype == 2:
                    v = cell.value
                    row.append(int(v) if v == int(v) else v)
                else:
                    row.append(cell.value)
            filas.append(tuple(row))
        return filas
    else:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))
        wb.close()
        return filas


def main():
    filepath = seleccionar_archivo()
    if not filepath:
        print("No se seleccionó ningún archivo.")
        sys.exit(1)

    print(f"\nArchivo: {os.path.basename(filepath)}")
    filas = leer_filas(filepath)
    print(f"Total filas: {len(filas)}")

    # ---- 1. VALORES ÚNICOS EN COLUMNA A (tipo de sección) ----
    col_a_vals = Counter()
    for row in filas:
        v = str(row[0]).strip().upper() if len(row) > 0 and row[0] else ""
        if v:
            col_a_vals[v] += 1

    print("\n" + "=" * 70)
    print("VALORES ÚNICOS EN COLUMNA A (tipo de fila)")
    print("=" * 70)
    for val, count in sorted(col_a_vals.items(), key=lambda x: -x[1]):
        print(f"  {count:>5}x  {val}")

    # ---- 2. COLUMNA B: clasificar por contexto ----
    # Buscar filas donde col C empieza con DETALLE -> es bloque nuevo con inst+dist
    # Buscar filas donde col A = HABERES -> nombre del empleado
    # Buscar filas donde col A está vacía y col B tiene datos -> metadatos (cargo, RD, uu, DNI, distrito)

    col_b_cargos = Counter()
    col_b_rds = Counter()
    col_b_uu = Counter()
    col_b_dnis = Counter()
    col_b_instituciones = Counter()
    col_b_distritos = Counter()
    col_b_otros = Counter()

    # Patrones
    import re
    rd_pat = re.compile(r'^(RD|RM|DS|LEY|DL|R\.D\.|R\.M\.|R\.N\.|R\.G\.)\w*\s*[\d\-./A-Za-z\s]+', re.IGNORECASE)
    uu_pat = re.compile(r'^uu-', re.IGNORECASE)
    dni_pat = re.compile(r'DNI\s*\d+', re.IGNORECASE)

    # Cargos conocidos
    cargos_known = [
        "PROF.", "PROF ", "PROF.DE AULA", "PROF.POR HORA",
        "DIRECTOR", "AUX.", "AUX.DE.EDUC.", "AUX.EDUCACION",
        "AUX.EDUC.", "AUX.BIBLIOT", "AUX.EDUCAC.",
        "ASESOR", "SECRETARIA", "SECRETARIO",
        "OFICINISTA", "ESPECIALISTA", "COORDINADOR",
        "JEFE", "SUBDIRECTOR", "PROMOTOR", "TECNICO",
        "PSICOLOGO", "ENFERMERO", "MEDICO", "TRABAJADOR",
        "TRAB.DE.SERV.", "TRAB.SERV", "TRAB.SERV.I", "TRAB.SERV.II",
        "ADMINISTRATIVO", "DOCENTE", "DOC.COORDINAD.",
        "INSPECTOR", "BIBLIOTECARIO", "LABORATORIO",
        "LABORAT.", "EDUCAC.", "EDUCACIÓN", "EDUCACION",
        "PER.SERV.II", "PERSONAL",
    ]
    distritos_known = [
        "ASIA", "CALANGO", "CERRO AZUL", "CHILCA", "COAYLLO", "IMPERIAL",
        "LUNAHUANÁ", "LUNAHUANA", "MALA", "NUEVO IMPERIAL", "PACARÁN",
        "QUILMANÁ", "QUILMANA", "SAN ANTONIO", "SAN LUIS",
        "SAN VICENTE DE CAÑETE", "SANTA CRUZ DE FLORES", "ZÚÑIGA", "ZUÑIGA",
    ]

    for row in filas:
        a = str(row[0]).strip().upper() if len(row) > 0 and row[0] else ""
        b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        c = str(row[2]).strip().upper() if len(row) > 2 and row[2] else ""

        if not b:
            continue

        # Si col C = DETALLE -> posible institución
        if "DETALLE" in c:
            col_b_instituciones[b.upper()] += 1
            continue

        # Si col A = HABERES -> nombre, no clasificar
        if a == "HABERES":
            continue

        # Clasificar col B según patrones
        b_upper = b.upper()

        if dni_pat.match(b):
            col_b_dnis[b] += 1
        elif rd_pat.match(b):
            col_b_rds[b] += 1
        elif uu_pat.match(b):
            col_b_uu[b] += 1
        elif b_upper in [d.upper() for d in distritos_known]:
            col_b_distritos[b] += 1
        elif any(b_upper.startswith(d.upper()) for d in distritos_known):
            col_b_distritos[b] += 1
        elif any(b_upper.startswith(c.upper()) for c in cargos_known):
            col_b_cargos[b] += 1
        else:
            # Check if it looks like a distrito but with variations
            is_dist = False
            b_norm = b_upper.replace("Ñ", "N").replace("É", "E").replace("Á", "A").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
            for d in distritos_known:
                d_norm = d.upper().replace("Ñ", "N").replace("É", "E").replace("Á", "A").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
                if d_norm in b_norm or b_norm in d_norm:
                    col_b_distritos[b] += 1
                    is_dist = True
                    break
            if not is_dist:
                # Check if it looks like institution with prefix
                inst_prefix = re.compile(r'^(CE\b|CB\b|C\.E\.|E\.E\.|I\.E\.|IE\b|CEBA|CEBE|EE\s|EE\b|JARDIN|COLEGIO|ESCUELA|INSTITUCION|INSTITUTO)', re.IGNORECASE)
                if inst_prefix.match(b):
                    col_b_instituciones[b.upper()] += 1
                else:
                    # Check if it could be a cargo not in our list
                    col_b_otros[b] += 1

    print("\n" + "=" * 70)
    print(f"VALORES ÚNICOS EN COLUMNA B - INSTITUCIONES ({len(col_b_instituciones)})")
    print("=" * 70)
    for val, count in sorted(col_b_instituciones.items(), key=lambda x: -x[1]):
        print(f"  {count:>5}x  {val}")

    print("\n" + "=" * 70)
    print(f"VALORES ÚNICOS EN COLUMNA B - DISTRITOS ({len(col_b_distritos)})")
    print("=" * 70)
    for val, count in sorted(col_b_distritos.items(), key=lambda x: -x[1]):
        print(f"  {count:>5}x  {val}")

    print("\n" + "=" * 70)
    print(f"VALORES ÚNICOS EN COLUMNA B - CARGOS ({len(col_b_cargos)})")
    print("=" * 70)
    for val, count in sorted(col_b_cargos.items(), key=lambda x: -x[1]):
        print(f"  {count:>5}x  {val}")

    print("\n" + "=" * 70)
    print(f"VALORES ÚNICOS EN COLUMNA B - RD / RESOLUCIONES ({len(col_b_rds)})")
    print("=" * 70)
    for val, count in sorted(col_b_rds.items(), key=lambda x: -x[1]):
        print(f"  {count:>5}x  {val}")

    print("\n" + "=" * 70)
    print(f"VALORES ÚNICOS EN COLUMNA B - UU / CÓDIGOS ({len(col_b_uu)})")
    print("=" * 70)
    for val, count in sorted(col_b_uu.items(), key=lambda x: -x[1]):
        print(f"  {count:>5}x  {val}")

    print("\n" + "=" * 70)
    print(f"VALORES ÚNICOS EN COLUMNA B - DNIs ({len(col_b_dnis)})")
    print("=" * 70)
    for val, count in sorted(col_b_dnis.items(), key=lambda x: -x[1]):
        print(f"  {count:>5}x  {val}")

    if col_b_otros:
        print("\n" + "=" * 70)
        print(f"VALORES ÚNICOS EN COLUMNA B - OTROS / NO CLASIFICADOS ({len(col_b_otros)})")
        print("=" * 70)
        for val, count in sorted(col_b_otros.items(), key=lambda x: -x[1]):
            print(f"  {count:>5}x  {val}")

    # ---- 3. COLUMNA C: valores únicos de conceptos ----
    col_c_conceptos = Counter()
    for row in filas:
        a = str(row[0]).strip().upper() if len(row) > 0 and row[0] else ""
        c = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        d = row[3] if len(row) > 3 else None
        if c and a not in ("HABERES", "DSCTOS", "TOTAL HABERES", "TOTAL DESCUENTOS", "TOTAL LIQUIDO", ""):
            col_c_conceptos[c.upper()] += 1

    print("\n" + "=" * 70)
    print(f"VALORES ÚNICOS EN COLUMNA C - CONCEPTOS ({len(col_c_conceptos)})")
    print("=" * 70)
    for val, count in sorted(col_c_conceptos.items(), key=lambda x: -x[1]):
        print(f"  {count:>5}x  {val}")

    print("\n" + "=" * 70)
    print("RESUMEN")
    print("=" * 70)
    print(f"  Instituciones únicas   : {len(col_b_instituciones)}")
    print(f"  Distritos únicos       : {len(col_b_distritos)}")
    print(f"  Cargos únicos          : {len(col_b_cargos)}")
    print(f"  RDs únicos             : {len(col_b_rds)}")
    print(f"  UUs únicos             : {len(col_b_uu)}")
    print(f"  DNIs únicos            : {len(col_b_dnis)}")
    print(f"  Conceptos únicos       : {len(col_c_conceptos)}")
    print(f"  No clasificados        : {len(col_b_otros)}")
    print()

    input("\nPresiona Enter para salir...")


if __name__ == "__main__":
    main()
